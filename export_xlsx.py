"""Excel export for Interval (backend-hardener split).

Moved verbatim from session_logger.py: duration options, filtering,
filename sanitizing, and the 4-sheet workbook builder. Requires openpyxl.
"""
import re
from collections import defaultdict
from datetime import datetime, timedelta, date

from timelib import parse_time

DURATION_OPTIONS = {
    "Today": ("today", None),
    "Last 3 days": ("days", 3),
    "Last 7 days": ("days", 7),
    "Last 12 days": ("days", 12),
    "Last 30 days": ("days", 30),
    "All time": ("all", None),
}


def filter_sessions(rows, category, tag, duration_label):
    mode, n = DURATION_OPTIONS.get(duration_label, ("all", None))
    now = datetime.now()
    start_of_today = datetime(now.year, now.month, now.day)
    out = []
    for r in rows:
        try:
            start = parse_time(r["start"])
        except Exception:
            continue
        if start is None:
            continue
        if mode == "today" and start < start_of_today:
            continue
        if mode == "days" and start < now - timedelta(days=n):
            continue
        if category != "All categories" and r.get("category", "") != category:
            continue
        if tag != "All tags" and r.get("tag", "") != tag:
            continue
        out.append(r)
    return out


def _sanitize_filename(s):
    s = str(s or "").strip()
    if not s or s in ("All categories", "All tags"):
        return "all"
    # filesystem-safe: replace invalid chars and collapse spaces
    s = re.sub(r'[\\/*?:"<>|]', "-", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_- ")
    return s[:48] or "all"


def _auto_export_name(category, tag, duration_label, rows):
    cat_safe = _sanitize_filename(category)
    tag_safe = _sanitize_filename(tag)
    today = date.today()
    end_str = today.strftime("%Y%m%d")
    mode, n = DURATION_OPTIONS.get(duration_label, ("all", None))
    if mode == "all":
        # start = earliest session start in filtered rows, else today
        starts = [parse_time(r.get("start")) for r in rows if parse_time(r.get("start"))]
        if starts:
            start_str = min(starts).date().strftime("%Y%m%d")
        else:
            start_str = end_str
    elif mode == "today":
        start_str = today.strftime("%Y%m%d")
    else:
        start_str = (datetime.now() - timedelta(days=n)).date().strftime("%Y%m%d")
    return f"{cat_safe}_{tag_safe}_{start_str}_{end_str}"


def _collision_free(out_dir, base):
    """Append _1, _2… when base exists. Returns a non-existing Path."""
    from pathlib import Path
    out_dir = Path(out_dir)
    path = out_dir / base
    if path.exists():
        stem = path.stem
        for i in range(1, 100):
            cand = out_dir / f"{stem}_{i}.xlsx"
            if not cand.exists():
                path = cand
                break
    return path


def export_name(category, tag, duration_label, rows, custom_name=None):
    """Resolve the export filename (without directory).

    Returns the file name including .xlsx. Collision handling is left to
    _collision_free so the API layer stays thin.
    """
    if custom_name and str(custom_name).strip():
        base = _sanitize_filename(str(custom_name).strip())
        base = re.sub(r"\.xlsx$", "", base, flags=re.I)
        base = _sanitize_filename(base) + ".xlsx"
        if base in (".xlsx", ""):
            base = _auto_export_name(category, tag, duration_label, rows) + ".xlsx"
        return base
    return _auto_export_name(category, tag, duration_label, rows) + ".xlsx"


def build_workbook(rows, keep_category_col):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions"

    header = ["Start", "End", "Duration (min)", "Doc (min)"]
    if keep_category_col:
        header.append("Category")
    header += ["Tag", "Sub-tag", "Describe", "Notes"]
    ws.append(header)

    hdr_fill = PatternFill("solid", fgColor="233447")
    hdr_font = Font(bold=True, color="EAF0F6", size=10)
    thin = Side(style="thin", color="2A3A50")
    hdr_border = Border(bottom=thin)

    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border

    # --- Sessions sheet (now includes Doc (min) per session so documentation tag/time is visible) ---
    for r in sorted(rows, key=lambda x: x.get("start", ""), reverse=True):
        try:
            dur = round((parse_time(r["end"]) - parse_time(r["start"])).total_seconds() / 60)
        except Exception:
            dur = 0
        doc_mins = round(int(r.get("doc_seconds") or 0) / 60, 1)
        if isinstance(doc_mins, float) and doc_mins.is_integer():
            doc_mins = int(doc_mins)
        row = [r["start"], r["end"], dur, doc_mins]
        if keep_category_col:
            row.append(r.get("category", ""))
        row += [r.get("tag", ""), r.get("sub_tag", ""),
                r.get("describe", ""), r.get("notes", "")]
        ws.append(row)

    widths = {"Start": 19, "End": 19, "Duration (min)": 13, "Doc (min)": 10, "Category": 18,
              "Tag": 18, "Sub-tag": 16, "Describe": 54, "Notes": 32}
    for idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header[idx - 1], 14)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if cell.column in (3, 4):  # Duration, Doc
                cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- By Category sheet (includes documentation time) ---
    ws2 = wb.create_sheet("By Category")
    header2 = ["Category", "Sessions", "Session time (min)", "Documentation (min)", "Total (min)", "Avg session (min)"]
    ws2.append(header2)
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="2B4A3A")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border

    # Aggregate per category from filtered rows (doc_seconds directly, so historic logs are included)
    from collections import Counter  # noqa: F401 (kept for parity with original imports)
    cats = sorted({r.get("category") or "Uncategorized" for r in rows})
    # also include cats that only have doc time but zero? already covered since doc is per session cat
    cat_stats = {}
    for cat in cats:
        cat_rows = [r for r in rows if (r.get("category") or "Uncategorized") == cat]
        n = len(cat_rows)
        sess_mins = 0
        doc_secs = 0
        for r in cat_rows:
            try:
                sess_mins += round((parse_time(r["end"]) - parse_time(r["start"])).total_seconds() / 60)
            except Exception:
                pass
            doc_secs += int(r.get("doc_seconds") or 0)
        doc_mins = round(doc_secs / 60)
        total = sess_mins + doc_mins
        avg = round(sess_mins / n, 1) if n else 0
        cat_stats[cat] = (n, sess_mins, doc_mins, total, avg)

    # sort by total desc
    for cat in sorted(cat_stats, key=lambda c: cat_stats[c][3], reverse=True):
        n, sess_mins, doc_mins, total, avg = cat_stats[cat]
        ws2.append([cat, n, sess_mins, doc_mins, total, avg])

    # total row
    if cat_stats:
        tot_n = sum(v[0] for v in cat_stats.values())
        tot_sess = sum(v[1] for v in cat_stats.values())
        tot_doc = sum(v[2] for v in cat_stats.values())
        tot_tot = sum(v[3] for v in cat_stats.values())
        tot_avg = round(tot_sess / tot_n, 1) if tot_n else 0
        ws2.append(["TOTAL", tot_n, tot_sess, tot_doc, tot_tot, tot_avg])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="233447")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = hdr_border

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            if cell.column == 1:
                cell.alignment = Alignment(vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    w2 = {"Category": 22, "Sessions": 12, "Session time (min)": 18, "Documentation (min)": 20, "Total (min)": 14, "Avg session (min)": 18}
    for idx, h in enumerate(header2, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w2.get(h, 14)
    ws2.freeze_panes = "A2"
    if ws2.max_row > 1:
        ws2.auto_filter.ref = ws2.dimensions
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    # --- By Tag sheet (better tags: per-tag breakdown incl. documentation) ---
    wsTag = wb.create_sheet("By Tag")
    headerTag = ["Tag", "Category", "Sessions", "Session time (min)", "Documentation (min)", "Total (min)"]
    wsTag.append(headerTag)
    for cell in wsTag[1]:
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="2F3A2B")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border
    # aggregate per (category,tag)
    from collections import defaultdict as _dd  # noqa: F401 (parity)
    tag_stats = {}
    for r in rows:
        cat = r.get("category") or "Uncategorized"
        tag = r.get("tag") or "—"
        key = (cat, tag)
        if key not in tag_stats:
            tag_stats[key] = {"n": 0, "sess": 0, "doc": 0}
        tag_stats[key]["n"] += 1
        try:
            tag_stats[key]["sess"] += round((parse_time(r["end"]) - parse_time(r["start"])).total_seconds() / 60)
        except Exception:
            pass
        tag_stats[key]["doc"] += int(r.get("doc_seconds") or 0)
    # convert doc secs to mins
    for k in list(tag_stats.keys()):
        tag_stats[k]["doc"] = round(tag_stats[k]["doc"] / 60)
        tag_stats[k]["total"] = tag_stats[k]["sess"] + tag_stats[k]["doc"]
    for (cat, tag), vals in sorted(tag_stats.items(), key=lambda x: -x[1]["total"]):
        wsTag.append([tag, cat, vals["n"], vals["sess"], vals["doc"], vals["total"]])
    if tag_stats:
        tot_n = sum(v["n"] for v in tag_stats.values())
        tot_sess = sum(v["sess"] for v in tag_stats.values())
        tot_doc = sum(v["doc"] for v in tag_stats.values())
        tot_tot = sum(v["total"] for v in tag_stats.values())
        wsTag.append(["TOTAL", "", tot_n, tot_sess, tot_doc, tot_tot])
        for cell in wsTag[wsTag.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="233447")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = hdr_border
    for row in wsTag.iter_rows(min_row=2, max_row=wsTag.max_row):
        for cell in row:
            if cell.column in (3, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
    wTag = {"Tag": 20, "Category": 20, "Sessions": 10, "Session time (min)": 17, "Documentation (min)": 19, "Total (min)": 13}
    for idx, h in enumerate(headerTag, 1):
        wsTag.column_dimensions[get_column_letter(idx)].width = wTag.get(h, 14)
    wsTag.freeze_panes = "A2"
    if wsTag.max_row > 1:
        wsTag.auto_filter.ref = wsTag.dimensions
    wsTag.sheet_properties.pageSetUpPr.fitToPage = True

    # --- Documentation Daily sheet (sum per day, broken down by category) ---
    ws3 = wb.create_sheet("Documentation daily")
    header3 = ["Date", "Category", "Doc (min)", "Sessions contributing"]
    ws3.append(header3)
    for cell in ws3[1]:
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="3E2F1A")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border

    # per day per category doc aggregation
    daily = defaultdict(lambda: defaultdict(int))
    daily_counts = defaultdict(lambda: defaultdict(int))
    for r in rows:
        ds = int(r.get("doc_seconds") or 0)
        if ds <= 0:
            continue
        dt = parse_time(r.get("start"))
        if dt is None:
            continue
        day = dt.date().isoformat()
        cat = r.get("category") or "Uncategorized"
        daily[day][cat] += ds
        daily_counts[day][cat] += 1

    # produce rows sorted by date desc, then cat total desc
    for day in sorted(daily.keys(), reverse=True):
        per_cat = daily[day]
        # sort cats in day by minutes desc
        for cat, secs in sorted(per_cat.items(), key=lambda x: -x[1]):
            mins = round(secs / 60, 1) if secs % 60 else round(secs / 60)
            # if integer, keep int
            if isinstance(mins, float) and mins.is_integer():
                mins = int(mins)
            ws3.append([day, cat, mins, daily_counts[day][cat]])

    if ws3.max_row == 1:
        ws3.append(["—", "No documentation time in this range.", "", ""])

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            if cell.column in (3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    w3 = {"Date": 14, "Category": 22, "Doc (min)": 12, "Sessions contributing": 22}
    for idx, h in enumerate(header3, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = w3.get(h, 14)
    ws3.freeze_panes = "A2"
    if ws3.max_row > 1:
        ws3.auto_filter.ref = ws3.dimensions
    ws3.sheet_properties.pageSetUpPr.fitToPage = True

    # print settings for all sheets
    for w in (ws, ws2, wsTag, ws3):
        w.sheet_properties.pageSetUpPr.fitToPage = True
        w.page_setup.orientation = "landscape"
        w.page_setup.fitToWidth = 1
        w.page_setup.fitToHeight = 0

    return wb
