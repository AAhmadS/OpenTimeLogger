import json
import os
import re
import subprocess
import sys
import threading
import uuid
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    import webview
except ImportError:
    webview = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import analytics as _analytics
except ImportError:
    _analytics = None

try:
    import ai as _ai
except ImportError:
    _ai = None

from ui import UI_HTML
from brand import AVATAR_DATA_URI

UI_HTML = UI_HTML.replace("const AVATAR_URI=null;",
                          'const AVATAR_URI="%s";' % AVATAR_DATA_URI)

APP_NAME = "OpenTimeLogger"

def app_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

BASE_DIR = app_dir()
DATA_FILE = BASE_DIR / "sessions.json"

FMT = "%Y-%m-%dT%H:%M:%S"
FMT_SHORT = "%Y-%m-%dT%H:%M"

DURATION_OPTIONS = {
    "Today": ("today", None),
    "Last 3 days": ("days", 3),
    "Last 7 days": ("days", 7),
    "Last 12 days": ("days", 12),
    "Last 30 days": ("days", 30),
    "All time": ("all", None),
}


def parse_time(v):
    if isinstance(v, datetime):
        return v
    if not isinstance(v, str):
        return None
    for f in (FMT, FMT_SHORT):
        try:
            return datetime.strptime(v, f)
        except ValueError:
            continue
    return None


def to_iso(dt):
    return dt.strftime(FMT)


def now_iso():
    return datetime.now().strftime(FMT)


def new_session(start_iso):
    return {"id": uuid.uuid4().hex, "start": start_iso, "end": None,
            "category": "", "tag": "", "sub_tag": "", "describe": "", "notes": "",
            "doc_seconds": 0}


def filter_sessions(rows, category, tag, duration_label):
    mode, n = DURATION_OPTIONS.get(duration_label, ("all", None))
    now = datetime.now()
    start_of_today = datetime(now.year, now.month, now.day)
    out = []
    for r in rows:
        try:
            start = parse_time(r["start"])
        except Exception:
            continue
        if start is None:
            continue
        if mode == "today" and start < start_of_today:
            continue
        if mode == "days" and start < now - timedelta(days=n):
            continue
        if category != "All categories" and r.get("category", "") != category:
            continue
        if tag != "All tags" and r.get("tag", "") != tag:
            continue
        out.append(r)
    return out


def _sanitize_filename(s):
    s = str(s or "").strip()
    if not s or s in ("All categories", "All tags"):
        return "all"
    # filesystem-safe: replace invalid chars and collapse spaces
    s = re.sub(r'[\\/*?:"<>|]', "-", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_- ")
    return s[:48] or "all"

def _auto_export_name(category, tag, duration_label, rows):
    cat_safe = _sanitize_filename(category)
    tag_safe = _sanitize_filename(tag)
    today = date.today()
    end_str = today.strftime("%Y%m%d")
    mode, n = DURATION_OPTIONS.get(duration_label, ("all", None))
    if mode == "all":
        # start = earliest session start in filtered rows, else today
        starts = [parse_time(r.get("start")) for r in rows if parse_time(r.get("start"))]
        if starts:
            start_str = min(starts).date().strftime("%Y%m%d")
        else:
            start_str = end_str
    elif mode == "today":
        start_str = today.strftime("%Y%m%d")
    else:
        start_str = (datetime.now() - timedelta(days=n)).date().strftime("%Y%m%d")
    return f"{cat_safe}_{tag_safe}_{start_str}_{end_str}"

def build_workbook(rows, keep_category_col):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions"

    header = ["Start", "End", "Duration (min)", "Doc (min)"]
    if keep_category_col:
        header.append("Category")
    header += ["Tag", "Sub-tag", "Describe", "Notes"]
    ws.append(header)

    hdr_fill = PatternFill("solid", fgColor="233447")
    hdr_font = Font(bold=True, color="EAF0F6", size=10)
    thin = Side(style="thin", color="2A3A50")
    hdr_border = Border(bottom=thin)

    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border

    # --- Sessions sheet (now includes Doc (min) per session so documentation tag/time is visible) ---
    for r in sorted(rows, key=lambda x: x.get("start", ""), reverse=True):
        try:
            dur = round((parse_time(r["end"]) - parse_time(r["start"])).total_seconds() / 60)
        except Exception:
            dur = 0
        doc_mins = round(int(r.get("doc_seconds") or 0) / 60, 1)
        if isinstance(doc_mins, float) and doc_mins.is_integer():
            doc_mins = int(doc_mins)
        row = [r["start"], r["end"], dur, doc_mins]
        if keep_category_col:
            row.append(r.get("category", ""))
        row += [r.get("tag", ""), r.get("sub_tag", ""),
                r.get("describe", ""), r.get("notes", "")]
        ws.append(row)

    widths = {"Start": 19, "End": 19, "Duration (min)": 13, "Doc (min)": 10, "Category": 18,
              "Tag": 18, "Sub-tag": 16, "Describe": 54, "Notes": 32}
    for idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header[idx - 1], 14)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if cell.column in (3, 4):  # Duration, Doc
                cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- By Category sheet (includes documentation time) ---
    ws2 = wb.create_sheet("By Category")
    header2 = ["Category", "Sessions", "Session time (min)", "Documentation (min)", "Total (min)", "Avg session (min)"]
    ws2.append(header2)
    for cell in ws2[1]:
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="2B4A3A")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border

    # Aggregate per category from filtered rows (doc_seconds directly, so historic logs are included)
    from collections import defaultdict, Counter
    cats = sorted({r.get("category") or "Uncategorized" for r in rows})
    # also include cats that only have doc time but zero? already covered since doc is per session cat
    cat_stats = {}
    for cat in cats:
        cat_rows = [r for r in rows if (r.get("category") or "Uncategorized") == cat]
        n = len(cat_rows)
        sess_mins = 0
        doc_secs = 0
        for r in cat_rows:
            try:
                sess_mins += round((parse_time(r["end"]) - parse_time(r["start"])).total_seconds() / 60)
            except Exception:
                pass
            doc_secs += int(r.get("doc_seconds") or 0)
        doc_mins = round(doc_secs / 60)
        total = sess_mins + doc_mins
        avg = round(sess_mins / n, 1) if n else 0
        cat_stats[cat] = (n, sess_mins, doc_mins, total, avg)

    # sort by total desc
    for cat in sorted(cat_stats, key=lambda c: cat_stats[c][3], reverse=True):
        n, sess_mins, doc_mins, total, avg = cat_stats[cat]
        ws2.append([cat, n, sess_mins, doc_mins, total, avg])

    # total row
    if cat_stats:
        tot_n = sum(v[0] for v in cat_stats.values())
        tot_sess = sum(v[1] for v in cat_stats.values())
        tot_doc = sum(v[2] for v in cat_stats.values())
        tot_tot = sum(v[3] for v in cat_stats.values())
        tot_avg = round(tot_sess / tot_n, 1) if tot_n else 0
        ws2.append(["TOTAL", tot_n, tot_sess, tot_doc, tot_tot, tot_avg])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="233447")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = hdr_border

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            if cell.column == 1:
                cell.alignment = Alignment(vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    w2 = {"Category": 22, "Sessions": 12, "Session time (min)": 18, "Documentation (min)": 20, "Total (min)": 14, "Avg session (min)": 18}
    for idx, h in enumerate(header2, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w2.get(h, 14)
    ws2.freeze_panes = "A2"
    if ws2.max_row > 1:
        ws2.auto_filter.ref = ws2.dimensions
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    # --- By Tag sheet (better tags: per-tag breakdown incl. documentation) ---
    wsTag = wb.create_sheet("By Tag")
    headerTag = ["Tag", "Category", "Sessions", "Session time (min)", "Documentation (min)", "Total (min)"]
    wsTag.append(headerTag)
    for cell in wsTag[1]:
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="2F3A2B")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border
    # aggregate per (category,tag)
    from collections import defaultdict as _dd
    tag_stats = {}
    for r in rows:
        cat = r.get("category") or "Uncategorized"
        tag = r.get("tag") or "—"
        key = (cat, tag)
        if key not in tag_stats:
            tag_stats[key] = {"n": 0, "sess": 0, "doc": 0}
        tag_stats[key]["n"] += 1
        try:
            tag_stats[key]["sess"] += round((parse_time(r["end"]) - parse_time(r["start"])).total_seconds() / 60)
        except Exception:
            pass
        tag_stats[key]["doc"] += int(r.get("doc_seconds") or 0)
    # convert doc secs to mins
    for k in list(tag_stats.keys()):
        tag_stats[k]["doc"] = round(tag_stats[k]["doc"] / 60)
        tag_stats[k]["total"] = tag_stats[k]["sess"] + tag_stats[k]["doc"]
    for (cat, tag), vals in sorted(tag_stats.items(), key=lambda x: -x[1]["total"]):
        wsTag.append([tag, cat, vals["n"], vals["sess"], vals["doc"], vals["total"]])
    if tag_stats:
        tot_n = sum(v["n"] for v in tag_stats.values())
        tot_sess = sum(v["sess"] for v in tag_stats.values())
        tot_doc = sum(v["doc"] for v in tag_stats.values())
        tot_tot = sum(v["total"] for v in tag_stats.values())
        wsTag.append(["TOTAL", "", tot_n, tot_sess, tot_doc, tot_tot])
        for cell in wsTag[wsTag.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="233447")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = hdr_border
    for row in wsTag.iter_rows(min_row=2, max_row=wsTag.max_row):
        for cell in row:
            if cell.column in (3, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
    wTag = {"Tag": 20, "Category": 20, "Sessions": 10, "Session time (min)": 17, "Documentation (min)": 19, "Total (min)": 13}
    for idx, h in enumerate(headerTag, 1):
        wsTag.column_dimensions[get_column_letter(idx)].width = wTag.get(h, 14)
    wsTag.freeze_panes = "A2"
    if wsTag.max_row > 1:
        wsTag.auto_filter.ref = wsTag.dimensions
    wsTag.sheet_properties.pageSetUpPr.fitToPage = True

    # --- Documentation Daily sheet (sum per day, broken down by category) ---
    ws3 = wb.create_sheet("Documentation daily")
    header3 = ["Date", "Category", "Doc (min)", "Sessions contributing"]
    ws3.append(header3)
    for cell in ws3[1]:
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="3E2F1A")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = hdr_border

    # per day per category doc aggregation
    daily = defaultdict(lambda: defaultdict(int))
    daily_counts = defaultdict(lambda: defaultdict(int))
    for r in rows:
        ds = int(r.get("doc_seconds") or 0)
        if ds <= 0:
            continue
        dt = parse_time(r.get("start"))
        if dt is None:
            continue
        day = dt.date().isoformat()
        cat = r.get("category") or "Uncategorized"
        daily[day][cat] += ds
        daily_counts[day][cat] += 1

    # produce rows sorted by date desc, then cat total desc
    for day in sorted(daily.keys(), reverse=True):
        per_cat = daily[day]
        # sort cats in day by minutes desc
        for cat, secs in sorted(per_cat.items(), key=lambda x: -x[1]):
            mins = round(secs / 60, 1) if secs % 60 else round(secs / 60)
            # if integer, keep int
            if isinstance(mins, float) and mins.is_integer():
                mins = int(mins)
            ws3.append([day, cat, mins, daily_counts[day][cat]])

    if ws3.max_row == 1:
        ws3.append(["—", "No documentation time in this range.", "", ""])

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            if cell.column in (3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    w3 = {"Date": 14, "Category": 22, "Doc (min)": 12, "Sessions contributing": 22}
    for idx, h in enumerate(header3, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = w3.get(h, 14)
    ws3.freeze_panes = "A2"
    if ws3.max_row > 1:
        ws3.auto_filter.ref = ws3.dimensions
    ws3.sheet_properties.pageSetUpPr.fitToPage = True

    # print settings for all sheets
    for w in (ws, ws2, wsTag, ws3):
        w.sheet_properties.pageSetUpPr.fitToPage = True
        w.page_setup.orientation = "landscape"
        w.page_setup.fitToWidth = 1
        w.page_setup.fitToHeight = 0

    return wb


class Store:
    def __init__(self):
        self._lock = threading.Lock()
        self.sessions = self._load()

    @staticmethod
    def _hydrate(r):
        s = new_session(r.get("start", ""))
        s["id"] = r.get("id", uuid.uuid4().hex)
        s["end"] = r.get("end")
        for k in ("category", "tag", "sub_tag", "describe", "notes"):
            s[k] = r.get(k, "")
        s["doc_seconds"] = r.get("doc_seconds", 0)
        for k in ("kind", "summary_of", "auto_describe", "summary_seconds"):
            if k in r:
                s[k] = r[k]
        return s

    def _load(self):
        raw = {}
        if DATA_FILE.exists():
            try:
                raw = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            except Exception:
                raw = {}
        out = []
        if "active" in raw:
            act = raw.get("active")
            if act and act.get("start"):
                out.append(new_session(act["start"]))
            for r in raw.get("sessions", []):
                out.append(self._hydrate(r))
        else:
            for r in raw.get("sessions", []):
                out.append(self._hydrate(r))
        out.sort(key=lambda x: x.get("start", ""), reverse=True)
        return out

    def _save(self):
        with self._lock:
            tmp = DATA_FILE.with_suffix(".json.tmp")
            tmp.write_text(json.dumps({"sessions": self.sessions},
                                      ensure_ascii=False, indent=2), encoding="utf-8")
            tmp.replace(DATA_FILE)

    def find(self, sid):
        for s in self.sessions:
            if s["id"] == sid:
                return s
        return None


class Api:
    def __init__(self):
        self.store = Store()
        self._last_rollover_day = None
        self._last_doc_total = None
        self._maybe_rollover()
        self._win = None

    def _ok(self):
        self._maybe_rollover()
        # store.sessions is kept sorted desc on every mutation, so no sort needed here
        # (previous version sorted on every call => O(n log n) on each tab visit)
        return {"sessions": self.store.sessions[:]}

    # ---------- window controls (frameless titlebar) ----------
    def win_minimize(self):
        if self._win:
            self._win.minimize()
        return {"ok": True}

    def win_maximize(self):
        if self._win:
            try:
                self._win.toggle_fullscreen()
            except Exception:
                self._win.maximize()
        return {"ok": True}

    def win_close(self):
        if self._win:
            self._win.destroy()
        return {"ok": True}

    def win_resize(self, width, height):
        if self._win and width and height:
            try:
                self._win.resize(int(width), int(height))
            except Exception:
                pass
        return {"ok": True}

    # ---------- analytics ----------
    def dashboard_stats(self, range_key="30d"):
        if _analytics is None:
            return {"error": "analytics module missing"}
        try:
            return _analytics.compute_dashboard(range_key)
        except Exception as e:
            return {"error": str(e)}

    # ---------- AI bridge (BYOK) ----------
    def ai_status(self):
        return {"available": _ai is not None}

    def ai_get_config(self):
        return _ai.load_config() if _ai else {"error": "ai module missing"}

    def ai_save_config(self, cfg):
        return _ai.save_config(cfg) if _ai else {"error": "ai module missing"}

    def ai_add_key(self, provider, label, key):
        return _ai.add_key(provider, label, key) if _ai else {"error": "ai module missing"}

    def ai_remove_key(self, key_id):
        return _ai.remove_key(key_id) if _ai else {"error": "ai module missing"}

    def ai_migrate_keys(self):
        return _ai.migrate_keys_to_keyring() if _ai else {"error": "ai module missing"}

    def ai_set_agent(self, agent_id, provider, key_id, model):
        return _ai.set_agent(agent_id, provider, key_id, model) if _ai else {"error": "ai module missing"}

    def ai_test_model(self, provider, key_id, model):
        return _ai.test_model(provider, key_id, model, "chat") if _ai else {"error": "ai module missing"}

    def ai_list_models(self, provider, task="chat"):
        return _ai.list_models(provider, task) if _ai else {"error": "ai module missing"}

    def ai_start_pipeline(self):
        return _ai.start_pipeline() if _ai else {"error": "ai module missing"}

    def ai_pipeline_status(self):
        return _ai.get_pipeline_status() if _ai else {"error": "ai module missing"}

    def ai_get_reports(self):
        return _ai.get_reports() if _ai else {"error": "ai module missing"}

    def ai_get_tasks(self):
        return _ai.get_tasks() if _ai else {"error": "ai module missing"}

    def ai_save_tasks(self, tasks):
        return _ai.save_tasks(tasks) if _ai else {"error": "ai module missing"}

    def ai_toggle_proposition(self, task_id, prop_id, accepted):
        return _ai.toggle_proposition(task_id, prop_id, accepted) if _ai else {"error": "ai module missing"}

    def ai_get_insights(self):
        return _ai.get_insights() if _ai else {"error": "ai module missing"}

    def ai_run_coach(self):
        return _ai.run_coach(_ai.load_config()) if _ai else {"error": "ai module missing"}

    def ai_set_consent(self, value):
        return _ai.set_consent(bool(value)) if _ai else {"error": "ai module missing"}

    def ai_export_dpo(self):
        return _ai.export_dpo_rows() if _ai else {"error": "ai module missing"}

    def ai_set_ideal_time(self, days, start, end):
        return _ai.set_ideal_time(days, start, end) if _ai else {"error": "ai module missing"}

    def ai_fallback(self, agent_id):
        if not _ai:
            return {"error": "ai module missing"}
        cfg = _ai.load_config()
        a = (cfg.get("agents") or {}).get(agent_id) or {}
        return _ai.fallback_model(agent_id, a.get("provider"), a.get("key_id"), a.get("model"))

    def ai_agents(self):
        return _ai.agents_catalog() if _ai else {"error": "ai module missing"}

    def ai_providers(self):
        return _ai.providers_catalog() if _ai else {"error": "ai module missing"}

    # ---------- ASR recording (Python-side mic capture) ----------
    def asr_begin(self):
        try:
            import sounddevice as sd
        except ImportError:
            return {"error": "sounddevice not installed"}
        if getattr(self, "_asr", None):
            return {"error": "already recording"}
        import audio_capture as _ac
        self._asr = _ac.Recorder()
        return self._asr.begin()

    def asr_stop(self, provider, key_id, model):
        rec = getattr(self, "_asr", None)
        if not rec:
            return {"error": "not recording"}
        self._asr = None
        audio = rec.stop()
        if not audio:
            return {"error": "no audio captured"}
        return _ai.transcribe(audio, provider, key_id, model) if _ai else {"error": "ai module missing"}

    def asr_state(self):
        rec = getattr(self, "_asr", None)
        if not rec:
            return {"recording": False, "seconds": 0}
        return {"recording": True, "seconds": rec.seconds()}

    def asr_transcribe(self, audio_b64, provider, key_id, model):
        return _ai.transcribe(audio_b64, provider, key_id, model) if _ai else {"error": "ai module missing"}

    def _maybe_rollover(self):
        """Log an end-of-day 'documentation' summary for every completed day
        that had documentation time, summing it up per category."""
        today = date.today()
        # performance throttle: skip heavy scan if day hasn't changed and total doc unchanged
        doc_total = sum(int(s.get("doc_seconds") or 0) for s in self.store.sessions if s.get("kind") != "daily-doc-summary")
        if self._last_rollover_day == today and self._last_doc_total == doc_total:
            return
        changed = False
        days = {}
        for s in self.store.sessions:
            if s.get("kind") == "daily-doc-summary":
                continue
            dt = parse_time(s.get("start"))
            if dt is None:
                continue
            days.setdefault(dt.date(), []).append(s)

        for day, sess in days.items():
            if day >= today:
                continue
            total = sum(int(s.get("doc_seconds") or 0) for s in sess)
            if total <= 0:
                continue
            per_cat = {}
            for s in sess:
                cat = s.get("category") or "Uncategorized"
                per_cat[cat] = per_cat.get(cat, 0) + int(s.get("doc_seconds") or 0)
            lines = ["%s: %d min" % (c, round(v / 60))
                     for c, v in sorted(per_cat.items(), key=lambda x: -x[1]) if round(v/60) > 0]
            if not lines:
                lines = ["%s: %d min" % (c, round(v / 60)) for c, v in sorted(per_cat.items(), key=lambda x: -x[1])]
            auto_describe = " | ".join(lines) + ("  ·  Total: %d min" % round(total / 60))

            existing = [s for s in self.store.sessions
                        if s.get("kind") == "daily-doc-summary"
                        and s.get("summary_of") == day.isoformat()]
            if existing:
                rec = existing[0]
                if rec.get("auto_describe") != auto_describe or rec.get("summary_seconds") != total:
                    if rec.get("describe", "") == rec.get("auto_describe", ""):
                        rec["describe"] = auto_describe
                    rec["auto_describe"] = auto_describe
                    rec["summary_seconds"] = total
                    changed = True
            else:
                key = "%sT23:59:59" % day.isoformat()
                rec = new_session(key)
                rec["end"] = key
                rec["category"] = "Documentation"
                rec["tag"] = "documentation"
                rec["describe"] = auto_describe
                rec["auto_describe"] = auto_describe
                rec["kind"] = "daily-doc-summary"
                rec["summary_of"] = day.isoformat()
                rec["summary_seconds"] = total
                self.store.sessions.append(rec)
                changed = True
        if changed:
            self.store._save()
        self._last_rollover_day = today
        self._last_doc_total = doc_total

    def _resolve_when(self, when):
        if not isinstance(when, dict):
            return None, "Invalid time option."
        t = when.get("type")
        if t == "now":
            return now_iso(), None
        if t == "at":
            dt = parse_time(when.get("value"))
            if dt is None:
                return None, "Please choose a valid time."
            return to_iso(dt), None
        return None, "Invalid time option."

    def _parse(self, v):
        dt = parse_time(v)
        if dt is None:
            return None, "Please choose a valid time."
        return to_iso(dt), None

    # ---------- API ----------
    def get_state(self):
        return self._ok()

    def list_exports(self):
        out_dir = DATA_FILE.parent / "exports"
        if not out_dir.exists():
            return {"exports": []}
        # include both legacy sessions_*.xlsx and new category_tag_date_date.xlsx
        files = sorted(out_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        exports = []
        for p in files[:80]:
            if p.name.startswith("~$"):
                continue
            try:
                st = p.stat()
                exports.append({
                    "name": p.name,
                    "path": str(p.resolve()),
                    "mtime": datetime.fromtimestamp(st.st_mtime).strftime(FMT),
                    "mtime_label": datetime.fromtimestamp(st.st_mtime).strftime("%d %b · %H:%M"),
                    "size_kb": round(st.st_size / 1024, 1)
                })
            except Exception:
                continue
        return {"exports": exports}

    def delete_export(self, path):
        if not path or not isinstance(path, str):
            return {"error": "Invalid path"}
        try:
            p = Path(path).resolve()
        except Exception:
            return {"error": "Invalid path"}
        out_dir = (DATA_FILE.parent / "exports").resolve()
        try:
            if out_dir not in p.parents and p.parent.resolve() != out_dir:
                if p.parent.resolve() != out_dir:
                    return {"error": "Access denied"}
        except Exception:
            return {"error": "Access denied"}
        if not p.exists():
            return {"error": "File not found"}
        if p.suffix.lower() != ".xlsx":
            return {"error": "Not an export file"}
        try:
            p.unlink()
            return {"ok": True}
        except Exception as e:
            return {"error": str(e)}

    def open_export(self, path):
        if not path or not isinstance(path, str):
            return {"error": "Invalid path"}
        try:
            p = Path(path).resolve()
        except Exception:
            return {"error": "Invalid path"}
        out_dir = (DATA_FILE.parent / "exports").resolve()
        # security: must be inside exports
        try:
            # allow exact file inside out_dir
            if out_dir not in p.parents and p.parent.resolve() != out_dir and p.resolve() != out_dir:
                # also handle case where p is direct child
                if p.parent.resolve() != out_dir:
                    return {"error": "Access denied"}
        except Exception:
            return {"error": "Access denied"}
        if not p.exists():
            return {"error": "File not found"}
        if p.suffix.lower() != ".xlsx":
            return {"error": "Not an export file"}
        try:
            if sys.platform == "win32":
                os.startfile(str(p))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(p)])
            else:
                subprocess.Popen(["xdg-open", str(p)])
            return {"ok": True, "path": str(p)}
        except Exception as e:
            return {"error": str(e)}

    def start_session(self, when, data=None):
        iso, err = self._resolve_when(when)
        if err:
            return {"error": err}
        s = new_session(iso)
        data = data or {}
        for k in ("category", "tag", "sub_tag", "describe", "notes"):
            s[k] = str(data.get(k, "")).strip()
        self.store.sessions.append(s)
        # keep sorted desc for fast _ok later (insert at correct position instead of resort)
        self.store.sessions.sort(key=lambda x: x.get("start", ""), reverse=True)
        self.store._save()
        return self._ok()

    def log_past_session(self, data):
        data = data or {}
        start_iso, err = self._parse(data.get("start"))
        if err:
            return {"error": err}
        end_iso, err = self._parse(data.get("end"))
        if err:
            return {"error": err}
        if end_iso < start_iso:
            return {"error": "End time must be after the start time."}
        s = new_session(start_iso)
        s["end"] = end_iso
        for k in ("category", "tag", "sub_tag", "describe", "notes"):
            s[k] = str(data.get(k, "")).strip()
        self.store.sessions.append(s)
        self.store.sessions.sort(key=lambda x: x.get("start", ""), reverse=True)
        self.store._save()
        return self._ok()

    def end_session(self, sid, when):
        s = self.store.find(sid)
        if not s:
            return {"error": "Session not found."}
        if s.get("end"):
            return {"error": "Session already ended."}
        iso, err = self._resolve_when(when)
        if err:
            return {"error": err}
        if iso < s["start"]:
            return {"error": "End time cannot be before the start time."}
        s["end"] = iso
        self.store._save()
        return self._ok()

    def update_session(self, sid, fields):
        s = self.store.find(sid)
        if not s:
            return {"error": "Session not found."}
        fields = fields or {}
        new_start, new_end = s["start"], s.get("end")
        if "start" in fields and fields.get("start"):
            iso, err = self._parse(fields["start"])
            if err:
                return {"error": err}
            new_start = iso
        if "end" in fields:
            if fields.get("end") in (None, ""):
                new_end = None
            else:
                iso, err = self._parse(fields["end"])
                if err:
                    return {"error": err}
                new_end = iso
        if new_end and new_end < new_start:
            return {"error": "End time cannot be before the start time."}

        new_vals = {}
        for k in ("category", "tag", "sub_tag", "describe", "notes"):
            if k in fields:
                new_vals[k] = str(fields.get(k) or "").strip()
        if "doc_seconds" in fields:
            try:
                new_vals["doc_seconds"] = max(0, int(fields.get("doc_seconds") or 0))
            except (TypeError, ValueError):
                pass

        eff = dict(s)
        eff.update(new_vals)
        if s.get("kind") != "daily-doc-summary" and new_end is not None:
            if not (eff["category"] and eff["tag"] and eff["describe"]):
                return {"error": "Category, tag and description are required. Sub-tag and notes are optional."}

        s["start"] = new_start
        s["end"] = new_end
        s.update(new_vals)
        self.store.sessions.sort(key=lambda x: x.get("start", ""), reverse=True)
        self.store._save()
        return self._ok()

    def reopen_session(self, sid):
        s = self.store.find(sid)
        if not s:
            return {"error": "Session not found."}
        s["end"] = None
        self.store._save()
        return self._ok()

    def delete_session(self, sid):
        s = self.store.find(sid)
        if not s:
            return {"error": "Session not found."}
        self.store.sessions.remove(s)
        self.store._save()
        return self._ok()

    def export_excel(self, category, tag, duration_label, custom_name=None):
        if openpyxl is None:
            return {"error": "openpyxl is not installed. Run: python -m pip install openpyxl"}
        rows = [s for s in self.store.sessions
                if s.get("end") and s.get("kind") != "daily-doc-summary"]
        rows = filter_sessions(rows, category, tag, duration_label)
        if not rows:
            return {"error": "No sessions match this filter."}
        keep_cat = category == "All categories"
        wb = build_workbook(rows, keep_cat)
        out_dir = DATA_FILE.parent / "exports"
        out_dir.mkdir(exist_ok=True)
        # naming: if custom_name provided, use it (sanitized), else auto category_tag_start_end
        if custom_name and str(custom_name).strip():
            base = _sanitize_filename(str(custom_name).strip())
            # preserve user extension handling; ensure .xlsx
            if not base.lower().endswith(".xlsx"):
                # remove trailing .xlsx if user included, then re-add sanitized
                base = re.sub(r"\.xlsx$", "", base, flags=re.I)
                base = _sanitize_filename(base) + ".xlsx"
            else:
                base = base  # already sanitized includes .xlsx? _sanitize would strip dot, so handle
                if not base.endswith(".xlsx"):
                    base += ".xlsx"
            # fallback if sanitize made empty
            if base in (".xlsx", ""):
                base = _auto_export_name(category, tag, duration_label, rows) + ".xlsx"
            path = out_dir / base
            # collision: if exists, append _1, _2
            if path.exists():
                stem = path.stem
                for i in range(1, 100):
                    cand = out_dir / f"{stem}_{i}.xlsx"
                    if not cand.exists():
                        path = cand
                        break
        else:
            auto = _auto_export_name(category, tag, duration_label, rows)
            path = out_dir / (auto + ".xlsx")
            if path.exists():
                stem = path.stem
                for i in range(1, 100):
                    cand = out_dir / f"{stem}_{i}.xlsx"
                    if not cand.exists():
                        path = cand
                        break
        wb.save(path)
        return {"path": str(path), "count": len(rows),
                "category_col": keep_cat, "name": path.name}


def main():
    if webview is None:
        print("pywebview is not installed. Run: python -m pip install pywebview")
        return
    api = Api()
    if _ai is not None:
        try:
            _ai.migrate_keys_to_keyring()  # best-effort: legacy inline -> locker
        except Exception:
            pass
    webview.settings["DRAG_REGION_SELECTOR"] = ".titlebar"
    webview.settings["DRAG_REGION_DIRECT_TARGET_ONLY"] = True
    window = webview.create_window(APP_NAME, html=UI_HTML, js_api=api,
                                   width=1180, height=780, min_size=(940, 640),
                                   frameless=True, easy_drag=False,
                                   background_color="#0a0f16")
    api._win = window
    window.events.closed += api.store._save
    webview.start()


if __name__ == "__main__":
    main()
